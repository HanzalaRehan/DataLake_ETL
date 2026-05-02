from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
import re
from typing import Optional

import openpyxl
import pandas as pd
import pdfplumber


ROOT = Path(__file__).resolve().parent
OUTPUT_CSV = ROOT / "mineral_production_all_years.csv"

MONTH_SEQUENCE = [
    "Jul",
    "Aug",
    "Sep",
    "Oct",
    "Nov",
    "Dec",
    "Jan",
    "Feb",
    "Mar",
    "Apr",
    "May",
    "Jun",
]

MONTH_PATTERN = re.compile(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\b", re.IGNORECASE)
YEAR_PATTERN = re.compile(r"(?P<start>\d{4})[-_/](?P<end>\d{2,4})")


@dataclass
class TableSpec:
    source_file: str
    source_kind: str
    source_sheet: str
    fiscal_year: Optional[str]
    header_row_index: int
    mineral_col: int
    province_col: int
    period_cols: list[int]
    total_cols: list[int]


def clean_text(value) -> Optional[str]:
    if value is None:
        return None
    text = str(value).replace("\xa0", " ").replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text or None


def normalize_header(text: Optional[str]) -> Optional[str]:
    cleaned = clean_text(text)
    if cleaned is None:
        return None
    cleaned = re.sub(r"\s*\d+\s*$", "", cleaned)
    cleaned = cleaned.replace("-", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned or None


def parse_number(value) -> Optional[float | int]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if isinstance(value, float) and value.is_integer():
            return int(value)
        return value
    text = clean_text(value)
    if text is None:
        return None
    if set(text) == {"#"}:
        return None
    if text in {"-", "--", "—"}:
        return None
    text = text.replace(",", "").replace(" ", "")
    if text in {"", "-"}:
        return None
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        numeric = float(text)
        return int(numeric) if numeric.is_integer() else numeric
    return None


def normalize_numeric_output(value):
    try:
        if value is None or pd.isna(value):
            return value
        numeric = float(value)
        if numeric.is_integer():
            return int(numeric)
        return numeric
    except Exception:
        return value


def extract_fiscal_year(text: str) -> Optional[str]:
    if not text:
        return None
    match = YEAR_PATTERN.search(text)
    if not match:
        return None
    start = int(match.group("start"))
    end_raw = match.group("end")
    if len(end_raw) == 2:
        end = (start // 100) * 100 + int(end_raw)
    else:
        end = int(end_raw)
    return f"{start}-{str(end)[-2:]}"


def fiscal_year_start(fiscal_year: Optional[str]) -> Optional[int]:
    if not fiscal_year:
        return None
    match = re.fullmatch(r"(\d{4})-(\d{2})", fiscal_year)
    if not match:
        return None
    return int(match.group(1))


def month_labels_for_year(fiscal_year: Optional[str]) -> list[str]:
    start = fiscal_year_start(fiscal_year)
    if start is None:
        return [f"Month {index + 1}" for index in range(12)]
    end_year = start + 1
    labels: list[str] = []
    for index, month in enumerate(MONTH_SEQUENCE):
        suffix = start % 100 if index < 6 else end_year % 100
        labels.append(f"{month}-{suffix:02d}")
    return labels


def is_total_label(label: Optional[str]) -> bool:
    if not label:
        return False
    lowered = label.lower()
    return "total" in lowered or "grand" in lowered or "year total" in lowered


def is_month_label(label: Optional[str]) -> bool:
    if not label:
        return False
    return bool(MONTH_PATTERN.search(label))


def row_texts(row) -> list[Optional[str]]:
    return [clean_text(value) for value in row]


def score_row_for_header(row) -> int:
    score = 0
    for text in row_texts(row):
        if not text:
            continue
        lowered = text.lower()
        if "mineral" in lowered:
            score += 6
        if "province" in lowered:
            score += 5
        if is_month_label(text):
            score += 2
        if re.fullmatch(r"q[1-4]", lowered):
            score -= 1
        if is_total_label(text):
            score += 1
        if "electricity" in lowered or "establishment" in lowered:
            score -= 6
    return score


def pick_best_sheet(workbook: openpyxl.Workbook):
    best_sheet = None
    best_row_index = 1
    best_score = -10**9
    for worksheet in workbook.worksheets:
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(10, worksheet.max_row), values_only=True),
            start=1,
        ):
            score = score_row_for_header(row)
            if score > best_score:
                best_score = score
                best_sheet = worksheet
                best_row_index = row_index
    if best_sheet is None:
        raise ValueError("No usable sheet found")
    return best_sheet, best_row_index


def pick_best_pdf_table(tables):
    best_table = None
    best_score = -10**9
    for table in tables:
        if not table:
            continue
        score = score_row_for_header(table[0])
        if score > best_score:
            best_score = score
            best_table = table
    if best_table is None:
        raise ValueError("No usable PDF table found")
    return best_table


def detect_table_spec(rows, source_file: str, source_kind: str, source_sheet: str, header_row_index: int | None = None) -> TableSpec:
    if not rows:
        raise ValueError("Empty table")

    if header_row_index is None:
        best_score = -10**9
        best_row_index = 0
        for idx, row in enumerate(rows[:10]):
            score = score_row_for_header(row)
            if score > best_score:
                best_score = score
                best_row_index = idx
        header_row_index = best_row_index

    header = row_texts(rows[header_row_index])

    mineral_col = None
    province_col = None
    for idx, label in enumerate(header):
        if not label:
            continue
        lowered = label.lower()
        if mineral_col is None and "name of mineral" in lowered:
            mineral_col = idx
        elif province_col is None and "province" in lowered:
            province_col = idx

    if mineral_col is None:
        mineral_col = 0
    if province_col is None:
        province_col = 1 if mineral_col == 0 and len(header) > 1 else min(mineral_col + 1, max(len(header) - 1, 0))

    fiscal_year = None
    for row in rows[:3]:
        for cell in row:
            fiscal_year = extract_fiscal_year(clean_text(cell) or "")
            if fiscal_year:
                break
        if fiscal_year:
            break
    if fiscal_year is None:
        fiscal_year = extract_fiscal_year(source_file)

    trailing_indices = list(range(province_col + 1, len(header)))
    total_cols = [index for index in trailing_indices if is_total_label(header[index])]
    period_cols = [index for index in trailing_indices if index not in total_cols]

    if len(period_cols) < 4:
        raise ValueError(
            f"Not enough monthly columns detected in {source_file} / {source_sheet}: {len(period_cols)}"
        )

    return TableSpec(
        source_file=source_file,
        source_kind=source_kind,
        source_sheet=source_sheet,
        fiscal_year=fiscal_year,
        header_row_index=header_row_index,
        mineral_col=mineral_col,
        province_col=province_col,
        period_cols=period_cols,
        total_cols=total_cols,
    )


def rows_to_records(rows, spec: TableSpec) -> list[dict[str, object]]:
    month_labels = month_labels_for_year(spec.fiscal_year)
    records: list[dict[str, object]] = []
    current_mineral: Optional[str] = None

    header_width = len(rows[spec.header_row_index])

    for row_index in range(spec.header_row_index + 1, len(rows)):
        row = list(rows[row_index])
        if len(row) < header_width:
            row.extend([None] * (header_width - len(row)))

        mineral = clean_text(row[spec.mineral_col]) if spec.mineral_col < len(row) else None
        province = clean_text(row[spec.province_col]) if spec.province_col < len(row) else None

        if mineral:
            current_mineral = mineral
        if current_mineral is None or not province:
            continue

        month_values = [parse_number(row[index]) if index < len(row) else None for index in spec.period_cols]
        total_values = [parse_number(row[index]) if index < len(row) else None for index in spec.total_cols]

        if all(value is None for value in month_values) and all(value is None for value in total_values):
            continue

        for period_order, (label, value) in enumerate(zip(month_labels, month_values), start=1):
            records.append(
                {
                    "source_file": spec.source_file,
                    "source_kind": spec.source_kind,
                    "source_sheet": spec.source_sheet,
                    "fiscal_year": spec.fiscal_year,
                    "mineral": current_mineral,
                    "province": province,
                    "period_type": "month",
                    "period_order": period_order,
                    "period": label,
                    "value": value,
                    "unit": "MT",
                }
            )

        for index, value in zip(spec.total_cols, total_values):
            records.append(
                {
                    "source_file": spec.source_file,
                    "source_kind": spec.source_kind,
                    "source_sheet": spec.source_sheet,
                    "fiscal_year": spec.fiscal_year,
                    "mineral": current_mineral,
                    "province": province,
                    "period_type": "year_total",
                    "period_order": 13,
                    "period": normalize_header(rows[spec.header_row_index][index]) or "Grand Total",
                    "value": value,
                    "unit": "MT",
                }
            )

    return records


def parse_workbook_file(path: Path) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(BytesIO(path.read_bytes()), data_only=True)
    sheet, header_row_index = pick_best_sheet(workbook)
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    spec = detect_table_spec(rows, path.name, "xlsx", sheet.title, header_row_index=header_row_index)
    return rows_to_records(rows, spec)


def parse_pdf_file(path: Path) -> list[dict[str, object]]:
    records: list[dict[str, object]] = []
    with pdfplumber.open(str(path)) as pdf:
        for page_number, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            if not tables:
                continue
            table = pick_best_pdf_table(tables)
            rows = [list(row) for row in table]
            try:
                spec = detect_table_spec(rows, path.name, "pdf", f"page-{page_number}")
            except ValueError:
                continue
            records.extend(rows_to_records(rows, spec))
    return records


def parse_source_file(path: Path) -> list[dict[str, object]]:
    with open(path, "rb") as handle:
        signature = handle.read(4)

    if signature == b"PK\x03\x04":
        return parse_workbook_file(path)
    if path.suffix.lower() == ".pdf":
        return parse_pdf_file(path)
    return []


def main() -> None:
    source_files = sorted(
        path
        for path in ROOT.iterdir()
        if path.is_file() and path.suffix.lower() in {".pdf", ".xlsx", ".xlsm", ".xls"}
    )

    records: list[dict[str, object]] = []
    for path in source_files:
        try:
            records.extend(parse_source_file(path))
        except Exception as exc:
            raise RuntimeError(f"Failed to parse {path.name}: {exc}") from exc

    frame = pd.DataFrame(records)
    if frame.empty:
        raise RuntimeError("No records were extracted from the source files")

    frame["value"] = pd.Series([normalize_numeric_output(value) for value in frame["value"]], dtype="object")
    frame["period_order"] = frame["period_order"].astype(int)

    sort_columns = [column for column in ["source_file", "source_sheet", "mineral", "province", "period_order", "period"] if column in frame.columns]
    frame = frame.sort_values(sort_columns, kind="stable").reset_index(drop=True)
    frame.to_csv(OUTPUT_CSV, index=False)

    print(f"Wrote {len(frame):,} rows to {OUTPUT_CSV}")
    print(frame.head(10).to_string(index=False))


if __name__ == "__main__":
    main()